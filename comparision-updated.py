import pandas as pd
import os
import openpyxl
from openpyxl.styles import PatternFill, Font
import boto3
from io import BytesIO

# Initialize the S3 client
s3 = boto3.client('s3')

# Define S3 bucket and output file key
bucket_name = 'github-csv'
output_key = 'comparision/TaxCal.xlsx'

# Get the current directory of the current subfolder
current_dir = os.path.dirname(os.path.abspath(__file__))
data3_path = os.path.join(current_dir, 'result1.jtl')

# Function to get the latest and second latest files based on prefix
def get_latest_files(prefix, count=2):
    response = s3.list_objects_v2(Bucket=bucket_name, Prefix=prefix)
    all_files = response.get('Contents', [])
    print(all_files)
    sorted_files = sorted(all_files, key=lambda x: x['LastModified'], reverse=True)
    return [file['Key'] for file in sorted_files[:count]]

# Function to get the prefix of the latest file
def get_latest_prefix():
    response = s3.list_objects_v2(Bucket=bucket_name)
    all_files = response.get('Contents', [])
    latest_file = max(all_files, key=lambda x: x['LastModified'])
    latest_key = latest_file['Key']
    prefix = os.path.dirname(latest_key) + 'Performance_test/'
    print(prefix)
    return prefix

# Get the prefix of the latest file
latest_prefix = get_latest_prefix()

# Get the latest and second latest files based on the latest prefix
latest_files = get_latest_files(latest_prefix)
data1_key = latest_files[0]
data2_key = latest_files[1]

# Read CSV files from S3
data1_obj = s3.get_object(Bucket=bucket_name, Key=data1_key)
data2_obj = s3.get_object(Bucket=bucket_name, Key=data2_key)

df1 = pd.read_csv(data1_obj['Body'])
df2 = pd.read_csv(data2_obj['Body'])
df_result = pd.read_csv(data3_path)

# Get the response count 500 for individual labels
response_500_counts = df_result[df_result['responseCode'] == '500'].groupby('label').size().reset_index(name='count_500')
response_501_counts = df_result[df_result['responseCode'] == '501'].groupby('label').size().reset_index(name='count_501')
response_502_counts = df_result[df_result['responseCode'] == '502'].groupby('label').size().reset_index(name='count_502')
response_503_counts = df_result[df_result['responseCode'] == '503'].groupby('label').size().reset_index(name='count_503')

# Create a DataFrame of all unique labels
all_labels = pd.DataFrame({'label': df_result['label'].unique()})

# Merge to include all labels, filling missing values with 0
all_labels_with_500 = pd.merge(all_labels, response_500_counts, on='label', how='left').fillna({'count_500': 0})
all_labels_with_501 = pd.merge(all_labels, response_501_counts, on='label', how='left').fillna({'count_501': 0})
all_labels_with_502 = pd.merge(all_labels, response_502_counts, on='label', how='left').fillna({'count_502': 0})
all_labels_with_503 = pd.merge(all_labels, response_503_counts, on='label', how='left').fillna({'count_503': 0})

new_df1 = pd.DataFrame({
    'Label': all_labels_with_500['label'],
    '500 Error Count': all_labels_with_500['count_500'],
    '501 Error Count': all_labels_with_501['count_501'],
    '502 Error Count': all_labels_with_502['count_502'],
    '503 Error Count': all_labels_with_503['count_503']
})

# Add the two values and store in a new column
new_df1['Sum of 5xx Error Count'] = new_df1['500 Error Count'] + new_df1['501 Error Count'] + new_df1['502 Error Count'] + new_df1['503 Error Count']

# Drop the '500 Error Count' column from the DataFrame
new_df1.drop(columns=['500 Error Count', '501 Error Count', '502 Error Count', '503 Error Count'], inplace=True)

#new_df1.columns = df1.columns.astype(str)

new_df1 = pd.merge(new_df1, df2[['Label', '90% Line']], on='Label', how='left', validate="many_to_one")
new_df1.rename(columns={'90% Line': '90% Line Previous Release'}, inplace=True)

new_df1 = pd.merge(new_df1, df1[['Label', '90% Line']], on='Label', how='left', validate="many_to_one")
new_df1.rename(columns={'90% Line': '90% Line Current Release'}, inplace=True)

new_df1['90% Line Difference'] = new_df1['90% Line Previous Release'] - new_df1['90% Line Current Release']
new_df1['% Response Time Deviation'] = (new_df1['90% Line Difference'] / new_df1['90% Line Current Release']) * 100

new_df1['Status'] = new_df1.apply(lambda row: 'PASS' if (row['Sum of 5xx Error Count'] == 0 and row['% Response Time Deviation'] >= -10) else 'FAIL', axis=1)

# Fetch Total Error % value
total_error_value = float(df1['Error %'].iloc[-1].replace('%', ''))

# Fetch Total 500/503 error count
error_count_sum = new_df1['Sum of 5xx Error Count'].sum()

status = 'PASS' if total_error_value <= 1 and error_count_sum <= 1 else 'FAIL'

new_df1 = new_df1.round(2)

# Save the DataFrame to an Excel file in memory
output = BytesIO()
new_df1.to_excel(output, index=True)
output.seek(0)

workbook = openpyxl.load_workbook(output)
sheet = workbook.active

# Define the fill color
fill_color_green = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
fill_color_red = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
fill_color_amber = PatternFill(start_color='FFBF00', end_color='FFBF00', fill_type='solid')

# Iterate through rows in a specific column (e.g., column A)
for row in range(2, sheet.max_row + 1):  # Assuming row 1 has headers
    cell = sheet[f'G{row}']  # Change 'A' to your specific column
    # Check your condition (e.g., cell value > 100)
    if cell.value < -10:
        cell.fill = fill_color_red
    elif cell.value >= -10 and cell.value < 0:
        cell.fill = fill_color_amber
    elif cell.value >= 0:
        cell.fill = fill_color_green

# Create a new column 'Final JMeter Test Result' and store the 'Status' value
sheet['I1'] = 'Final JMeter Test Result'
sheet['I1'].font = Font(bold=True)
sheet['I2'] = status
sheet['I2'].font = Font(bold=True)

# Apply formatting to the 'Status' cell based on the value
if status == 'FAIL':
    sheet['I2'].fill = fill_color_red
else:
    sheet['I2'].fill = fill_color_green

# Save the changes to the workbook in memory
output = BytesIO()
workbook.save(output)
output.seek(0)

# Upload the Excel file to S3
s3.put_object(Bucket=bucket_name, Key=output_key, Body=output.getvalue())